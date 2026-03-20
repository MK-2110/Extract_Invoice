#!/usr/bin/env python3
"""
================================================================
 Indian Customs Shipping Bill  -  Data Extractor
 No API / No external service - pure PDF parsing
================================================================
 Install  :  pip install pdfplumber pandas openpyxl

 Run (folder) :  python extract_shipping_bill.py <folder_path>
 Run (single) :  python extract_shipping_bill.py <file.pdf>

 Output files (saved beside the PDFs):
   <name>_extracted.json
   all_shipping_bills_extracted.json
   all_shipping_bills_extracted.xlsx
================================================================
"""

import re
import sys
import json
import traceback
from pathlib import Path
from typing import Optional

import pdfplumber          # type: ignore[import]
import pandas as pd        # type: ignore[import]
from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import]


# ----------------------------------------------------------------
#  PDF TABLE EXTRACTION SETTINGS
# ----------------------------------------------------------------
TABLE_SETTINGS: dict = {
    "vertical_strategy":    "lines",
    "horizontal_strategy":  "lines",
    "snap_tolerance":       4,
    "join_tolerance":       4,
    "edge_min_length":      5,
    "min_words_vertical":   1,
    "min_words_horizontal": 1,
    "text_tolerance":       4,
}

# Column order matching the reference Excel format
COLUMN_ORDER: list = [
    "SB_No", "SB_Date", "Invoice_No",
    "Port Code", "Consignee Name", "G.WT", "PKG",
    "Description", "Quantity", "UQC", "Rate",
    "Value (F/C)", "Exchange Rate",
    "Source_File",
]


# ================================================================
#  STEP 1 - READ PDF  (text + tables)
# ================================================================

def read_pdf(pdf_path: str):
    """
    Returns
    -------
    full_text : str   all pages joined as one string
    tables    : list  [{page: int, rows: list[list[str]]}]
    """
    texts:  list = []
    tables: list = []

    with pdfplumber.open(pdf_path) as pdf:
        for pg_num, page in enumerate(pdf.pages):

            # ── plain text ────────────────────────────────────────
            texts.append(
                page.extract_text(x_tolerance=3, y_tolerance=3) or ""
            )

            # ── tables ────────────────────────────────────────────
            raw_list = page.extract_tables(TABLE_SETTINGS) or []
            for raw in raw_list:
                rows: list = []
                for row in (raw or []):
                    if not isinstance(row, list):
                        continue
                    # FIX: explicit str() conversion handles None cells
                    cleaned: list = [
                        str(c).strip().replace("\n", " ") if c is not None else ""
                        for c in row
                    ]
                    if any(cleaned):
                        rows.append(cleaned)
                if rows:
                    tables.append({"page": pg_num, "rows": rows})

    return "\n".join(texts), tables


# ================================================================
#  STEP 2 - FIELD PARSERS
# ================================================================

# ----------------------------------------------------------------
#  2a.  Port Code / SB No / SB Date
#  Text always contains:  "INMAA4  5376650  17-SEP-25"
# ----------------------------------------------------------------

def parse_port_sb(text: str):
    pattern = (
        r"\b([A-Z]{2}[A-Z0-9]{2,6})"
        r"\s+(\d{5,8})"
        r"\s+(\d{2}-(?:JAN|FEB|MAR|APR|MAY|JUN"
        r"|JUL|AUG|SEP|OCT|NOV|DEC)-\d{2,4})\b"
    )
    m = re.search(pattern, text, re.IGNORECASE)
    if m:
        return m.group(1), m.group(2), m.group(3).upper()
    return None, None, None


# ----------------------------------------------------------------
#  2b.  PKG and G.WT
#  Text line:  "PKG 63  G.WT  KGS  750"
# ----------------------------------------------------------------

def parse_pkg_gwt(text: str):
    pkg: Optional[str] = None
    gwt: Optional[str] = None

    m = re.search(r"\bPKG\s+(\d+)", text, re.IGNORECASE)
    if m:
        pkg = m.group(1)

    m = re.search(r"G\.WT\s+(?:KGS\s+)?(\d+(?:\.\d+)?)", text, re.IGNORECASE)
    if m:
        gwt = m.group(1)

    return pkg, gwt


# ----------------------------------------------------------------
#  2c.  Invoice Number   e.g.  JT-138/25-26
# ----------------------------------------------------------------

def parse_invoice_no(text: str) -> Optional[str]:
    m = re.search(r"\b([A-Z]{1,6}-\d+/\d{2,4}-\d{2,4})\b", text)
    if m:
        return m.group(1)
    m = re.search(r"\b([A-Z]{1,6}[-/]\d[A-Z0-9\-/]{3,15})\b", text)
    if m:
        return m.group(1)
    return None


# ----------------------------------------------------------------
#  2d.  Exchange Rate + Foreign Currency Code
#  Text pattern: "1 SGD INR 67.2"  -> rate "67.2", currency "SGD"
# ----------------------------------------------------------------

def parse_exchange_rate(text: str) -> Optional[str]:
    m = re.search(r"\b1\s+([A-Z]{3})\s+INR\s+([\d.]+)\b", text)
    if m:
        return m.group(2).strip()
    return None


def parse_currency(text: str) -> Optional[str]:
    """Return the foreign currency code, e.g. 'SGD', 'AED', 'USD'."""
    m = re.search(r"\b1\s+([A-Z]{3})\s+INR\s+[\d.]+\b", text)
    if m:
        return m.group(1).upper()
    return None


# ----------------------------------------------------------------
#  2e.  Consignee Name
#
#  Table page 0:
#    label row → "7.CONSIGNEE NAME & ADDRESS"  at column index j
#    value row → "P JFA FLOWERS & TRADING PTE.LTD"  at same index j
#
#  FIX: use positive  `if i + 1 < len(rows):`  so Pylance is happy
# ----------------------------------------------------------------

def parse_consignee(tables: list) -> Optional[str]:
    for tbl in tables:
        rows: list = tbl["rows"]
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                if not re.search(r"CONSIGNEE\s+NAME", cell, re.IGNORECASE):
                    continue

                # FIX: positive guard instead of `if i+1 >= len: continue`
                if i + 1 < len(rows):
                    next_row: list = rows[i + 1]

                    # Primary: same column index as the label
                    val: str = next_row[j].strip() if j < len(next_row) else ""

                    # Fallback: scan next row for a company-like value
                    if len(val) < 4:
                        for v in next_row:
                            v = v.strip()
                            if (len(v) > 4
                                    and "EXPORTER" not in v.upper()
                                    and "SAME AS"  not in v.upper()
                                    and not re.match(r"^\d+\.", v)):
                                val = v
                                break

                    if len(val) > 3:
                        # Strip leading single-letter artifact "P JFA..." -> "JFA..."
                        cleaned: str = re.sub(r"^[A-Z]\s+", "", val)
                        return cleaned if len(cleaned) > 3 else val

    return None


# ----------------------------------------------------------------
#  2f.  Item Details  (Description, Qty, UQC, Rate, Value F/C)
#
#  Header row (many empty cells from merged-cell PDF grid):
#    idx 0  : rotated-text artifact  → skip
#    idx 1  : '1.ItemSNo'
#    idx 4  : 'T 3.DESCRIPTION'      → DESCRIPTION
#    idx 10 : '4.QUANTITY'
#    idx 13 : '5.UQC'
#    idx 16 : '6.RATE'
#    idx 18 : '7.VALUE(F/C)'
#  Data rows share the same column indices as the header.
# ----------------------------------------------------------------

def _get_cell(row: list, cols: dict, key: str) -> str:
    """Read one cell value using a pre-built column-index map."""
    idx = cols.get(key)
    if idx is not None and idx < len(row):
        return str(row[idx]).strip()
    return ""


def _map_columns(hdr: list):
    """Build {field -> column_index} from the header row."""
    cols:    dict          = {}
    sno_idx: Optional[int] = None

    for j, h in enumerate(hdr):
        hu = str(h).upper()
        if "DESCRIPTION" in hu:
            cols["desc"] = j
        elif "QUANTITY" in hu:
            cols["qty"]  = j
        elif "UQC" in hu:
            cols["uqc"]  = j
        elif re.search(r"\bRATE\b", hu) and "EXCHANGE" not in hu:
            cols["rate"] = j
        elif "VALUE" in hu and ("F/C" in hu or "VALUE(F" in hu):
            cols["val"]  = j
        elif re.search(r"ITEM\s*S[N]?O|1\.ITEM", hu):
            sno_idx = j

    return cols, sno_idx


def _is_data_row(row: list, sno_idx: Optional[int]) -> bool:
    """True when a row contains a serial number (1, 2, 3 …)."""
    if sno_idx is not None and sno_idx < len(row):
        if re.fullmatch(r"\d{1,3}", str(row[sno_idx]).strip()):
            return True
    for cell in row[:6]:
        if re.fullmatch(r"\d{1,2}", str(cell).strip()):
            return True
    return False


def parse_items(tables: list) -> list:
    items: list = []
    seen:  set  = set()

    for tbl in tables:
        rows: list = tbl["rows"]
        flat: str  = " ".join(" ".join(str(c) for c in r) for r in rows).upper()

        if "DESCRIPTION" not in flat or "QUANTITY" not in flat:
            continue

        # Find header row
        hdr_idx: Optional[int] = None
        for i, row in enumerate(rows):
            rf = " ".join(str(c) for c in row).upper()
            if "DESCRIPTION" in rf and "QUANTITY" in rf and "RATE" in rf and "VALUE" in rf:
                hdr_idx = i
                break

        if hdr_idx is None:
            continue

        cols, sno_idx = _map_columns(rows[hdr_idx])

        if "desc" not in cols:
            continue

        for row in rows[hdr_idx + 1:]:
            if not row:
                continue
            if not _is_data_row(row, sno_idx):
                continue

            desc: str = _get_cell(row, cols, "desc")
            qty:  str = _get_cell(row, cols, "qty")

            if len(desc) < 2:
                continue

            key_pair = (desc.upper(), qty)
            if key_pair in seen:
                continue
            seen.add(key_pair)

            items.append({
                "Description": desc,
                "Quantity":    qty,
                "UQC":         _get_cell(row, cols, "uqc"),
                "Rate":        _get_cell(row, cols, "rate"),
                "Value (F/C)": _get_cell(row, cols, "val"),
            })

    return items


# ================================================================
#  STEP 3 - MAIN EXTRACTOR
# ================================================================

def parse_shipping_bill(pdf_path: str) -> list:
    """
    Parse one PDF.
    Returns list of dicts, one per item line.
    Header fields repeat on every item row.
    """
    print("    Reading :", Path(pdf_path).name)

    text, tables = read_pdf(pdf_path)

    port_code, sb_no, sb_date = parse_port_sb(text)
    pkg, gwt                  = parse_pkg_gwt(text)
    invoice_no                = parse_invoice_no(text)
    exchange_rate             = parse_exchange_rate(text)
    currency                  = parse_currency(text)          # e.g. "SGD"
    consignee                 = parse_consignee(tables)
    items                     = parse_items(tables)

    # Prefix Rate with currency symbol (e.g. "SGD 4")
    if currency:
        for item in items:
            raw_rate: str = item.get("Rate", "")
            if raw_rate and not raw_rate.upper().startswith(currency):
                item["Rate"] = f"{currency} {raw_rate}"

    header: dict = {
        "SB_No":          sb_no,
        "SB_Date":        sb_date,
        "Invoice_No":     invoice_no,
        "Port Code":      port_code,
        "Consignee Name": consignee,
        "G.WT":           gwt,
        "PKG":            pkg,
        "Exchange Rate":  exchange_rate,
    }

    if not items:
        return [header]

    return [{**header, **item} for item in items]


# ================================================================
#  STEP 4 - EXCEL WRITER
# ================================================================

def save_excel(all_rows: list, excel_path: Path) -> None:
    df = pd.DataFrame(all_rows)

    for col in COLUMN_ORDER:
        if col not in df.columns:
            df[col] = None
    df = df[COLUMN_ORDER]

    with pd.ExcelWriter(str(excel_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Shipping Bills")
        ws = writer.sheets["Shipping Bills"]

        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        alt_fill = PatternFill("solid", fgColor="EBF3FB")
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill

        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=8,
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

        ws.freeze_panes = "A2"


# ================================================================
#  STEP 5 - PROCESS FOLDER / SINGLE FILE
# ================================================================

def process(input_path: str, output_dir: Optional[str] = None) -> None:
    inp = Path(input_path)

    if inp.is_dir():
        pdf_files = sorted(inp.glob("*.pdf"))
        out_dir   = Path(output_dir) if output_dir else inp
    elif inp.is_file() and inp.suffix.lower() == ".pdf":
        pdf_files = [inp]
        out_dir   = Path(output_dir) if output_dir else inp.parent
    else:
        print("ERROR: provide a .pdf file or a folder containing PDFs.")
        sys.exit(1)

    if not pdf_files:
        print("No PDF files found in:", input_path)
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    all_rows: list = []

    for pdf_path in pdf_files:
        print("\n" + "-" * 55)
        print("  File :", pdf_path.name)
        try:
            rows = parse_shipping_bill(str(pdf_path))
            for r in rows:
                r["Source_File"] = pdf_path.name
            all_rows.extend(rows)

            per_json = out_dir / (pdf_path.stem + "_extracted.json")
            per_json.write_text(
                json.dumps(rows, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
            print("    Rows :", len(rows))
            print("    JSON :", per_json.name)

            for r in rows:
                print("    SB=%-8s Date=%-12s Inv=%-16s Consignee=%s" % (
                    r.get("SB_No", ""), r.get("SB_Date", ""),
                    r.get("Invoice_No", ""), r.get("Consignee Name", "")))
                print("    Desc=%-35s Qty=%-5s UQC=%-5s Rate=%-5s Val=%-8s ExRate=%s" % (
                    r.get("Description", ""), r.get("Quantity", ""),
                    r.get("UQC", ""), r.get("Rate", ""),
                    r.get("Value (F/C)", ""), r.get("Exchange Rate", "")))

        except Exception as exc:
            print("    ERROR:", exc)
            traceback.print_exc()

    if not all_rows:
        print("\nNo data extracted.")
        return

    comb_json = out_dir / "all_shipping_bills_extracted.json"
    comb_json.write_text(
        json.dumps(all_rows, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    excel_path = out_dir / "all_shipping_bills_extracted.xlsx"
    save_excel(all_rows, excel_path)

    print("\n" + "=" * 55)
    print("  Combined JSON :", comb_json.name)
    print("  Excel         :", excel_path.name)
    print("  Total rows    :", len(all_rows))
    print("=" * 55)


# ================================================================
#  ENTRY POINT
# ================================================================

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python extract_shipping_bill.py <pdf_or_folder> [output_dir]")
        print()
        print("Examples:")
        print("  python extract_shipping_bill.py C:/Users/user/Desktop/Invoice_Extract/")
        print("  python extract_shipping_bill.py C:/Users/user/Desktop/Invoice_Extract/True__1_.pdf")
        sys.exit(0)

    process(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)

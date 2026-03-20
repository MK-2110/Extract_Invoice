"""
================================================================
 Shipping Bill Extractor  -  Streamlit UI
 Run:  streamlit run streamlit_app.py
================================================================
 All uploaded PDFs and generated Excel files are stored only
 in Python's tempfile (OS temp folder) and are NEVER written
 to the project folder.  Temp files are deleted automatically
 when the session ends or the browser tab is closed.
================================================================
"""

import io
import tempfile
import traceback
from pathlib import Path

import pandas as pd                          # type: ignore[import]
import streamlit as st                       # type: ignore[import]
from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import]

from extract_shipping_bill import parse_shipping_bill       # type: ignore[import]

# ── Column order ──────────────────────────────────────────────
COLUMN_ORDER = [
    "SB_No", "SB_Date", "Invoice_No",
    "Port Code", "Consignee Name", "G.WT", "PKG",
    "Description", "Quantity", "UQC", "Rate",
    "Value (F/C)", "Exchange Rate",
    "Source_File",
]


# ================================================================
#  Excel builder  (returns bytes — no file written to disk)
# ================================================================

def build_excel_bytes(all_rows: list) -> bytes:
    df = pd.DataFrame(all_rows)
    for col in COLUMN_ORDER:
        if col not in df.columns:
            df[col] = None
    df = df[COLUMN_ORDER]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Shipping Bills")
        ws = writer.sheets["Shipping Bills"]

        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        alt_fill = PatternFill("solid", fgColor="EBF3FB")
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill

        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=8,
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

        ws.freeze_panes = "A2"

    return buf.getvalue()


# ================================================================
#  Page config
# ================================================================

st.set_page_config(
    page_title="Shipping Bill Extractor",
    page_icon="📄",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ────────────────────────────────────────────────
st.markdown("""
<style>
  /* ── global ─────────────────────────────────────── */
  html, body, [data-testid="stAppViewContainer"] {
    background: #f8fafc !important; /* light gray/blue background */
    color: #0f172a !important;      /* dark slate text */
  }
  [data-testid="stHeader"] { background: transparent !important; }
  section[data-testid="stMain"] > div { padding-top: 0 !important; }

  /* Ensure text elements have good contrast */
  .stMarkdown, .stMarkdown p, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4, .stMarkdown h5, .stMarkdown h6, .stMarkdown li {
      color: #0f172a !important;
  }
  
  /* Metric labels and values */
  [data-testid="stMetricLabel"] { color: #475569 !important; }
  [data-testid="stMetricValue"] { color: #0f172a !important; }

  /* ── header banner ───────────────────────────────── */
  .app-header {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 14px;
    padding: 22px 28px;
    margin-bottom: 24px;
    display: flex; align-items: center; gap: 16px;
    box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05), 0 2px 4px -1px rgba(0,0,0,0.03);
  }
  .app-header .logo {
    width:48px; height:48px; border-radius:12px;
    background: linear-gradient(135deg,#3b82f6,#6366f1);
    display:flex; align-items:center; justify-content:center;
    font-size:24px; flex-shrink:0;
    color: #ffffff;
  }
  .app-header h1 { font-size:1.4rem; font-weight:700; color:#0f172a !important; margin:0; }
  .app-header p  { font-size:.82rem; color:#475569 !important; margin:0; }

  /* ── metric cards ────────────────────────────────── */
  [data-testid="metric-container"] {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 10px !important;
    padding: 16px !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important;
  }

  /* ── file uploader ───────────────────────────────── */
  [data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important;
    border: 2px dashed #cbd5e1 !important;
    border-radius: 12px !important;
  }
  [data-testid="stFileUploaderDropzone"]:hover {
    border-color: #3b82f6 !important;
    background: #eff6ff !important;
  }
  [data-testid="stFileUploaderDropzone"] section { color: #334155 !important; }
  [data-testid="stFileUploaderDropzone"] small { color: #64748b !important; }

  /* ── buttons ─────────────────────────────────────── */
  .stButton > button[kind="primary"] {
    background: linear-gradient(135deg,#3b82f6,#6366f1) !important;
    color: #ffffff !important;
    border: none !important;
    box-shadow: 0 4px 14px rgba(59,130,246,0.3) !important;
  }
  .stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg,#2563eb,#4f46e5) !important;
  }

  /* ── success / warning / error ───────────────────── */
  [data-testid="stAlert"] { border-radius: 10px !important; }

  /* ── progress bar ────────────────────────────────── */
  [data-testid="stProgressBar"] > div > div {
    background: linear-gradient(90deg,#3b82f6,#6366f1) !important;
    border-radius: 99px !important;
  }

  /* ── dataframe ───────────────────────────────────── */
  [data-testid="stDataFrame"] { border-radius: 10px !important; }

  /* ── download button ─────────────────────────────── */
  [data-testid="stDownloadButton"] > button {
    width: 100% !important;
    background: linear-gradient(135deg,#16a34a,#15803d) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 10px !important;
    font-size: 1rem !important;
    font-weight: 700 !important;
    padding: 14px !important;
    box-shadow: 0 4px 20px rgba(22,163,74,.3) !important;
  }
  [data-testid="stDownloadButton"] > button:hover {
    background: linear-gradient(135deg,#15803d,#166534) !important;
    box-shadow: 0 6px 20px rgba(22,163,74,.4) !important;
  }

  /* ── divider ─────────────────────────────────────── */
  hr { border-color: #cbd5e1 !important; }
</style>
""", unsafe_allow_html=True)


# ================================================================
#  Header
# ================================================================

st.markdown("""
<div class="app-header">
  <div class="logo">📄</div>
  <div>
    <h1>Shipping Bill Extractor</h1>
    <p>Indian Customs EDI &nbsp;·&nbsp; Bulk PDF Data Extraction &nbsp;·&nbsp; No API Required</p>
  </div>
</div>
""", unsafe_allow_html=True)


# ================================================================
#  File uploader
# ================================================================

st.markdown("#### 📂 Upload PDF Documents")
uploaded = st.file_uploader(
    label="Drag & drop or click to browse — PDF only",
    type=["pdf"],
    accept_multiple_files=True,
    help="Upload Indian Customs Shipping Bill PDFs. Minimum 10 documents recommended.",
)

# Warning if < 10
if uploaded:
    if len(uploaded) < 10:
        st.warning(
            f"⚠️  **{len(uploaded)} file(s) selected.** "
            "For bulk extraction, uploading at least **10 documents** is recommended.",
            icon="⚠️",
        )
    else:
        st.success(f"✅  **{len(uploaded)} PDF(s) ready** for extraction.", icon="✅")

st.markdown("---")

# ================================================================
#  Extract button + processing
# ================================================================

if uploaded:
    if st.button("⚡  Extract All Documents", use_container_width=True, type="primary"):

        all_rows:    list  = []
        file_errors: list  = []

        # ── Progress UI ───────────────────────────────────────
        progress_bar  = st.progress(0, text="Preparing…")
        status_text   = st.empty()
        log_container = st.container()

        total = len(uploaded)

        for idx, uf in enumerate(uploaded, 1):
            fname = uf.name
            status_text.markdown(
                f"**Processing** `{fname}` &nbsp;({idx} / {total})"
            )
            progress_bar.progress(
                (idx - 1) / total,
                text=f"Extracting {idx}/{total} — {fname}",
            )

            # ── Write to temp file, manually delete after (Windows-safe) ──
            try:
                tmp = tempfile.NamedTemporaryFile(
                    suffix=".pdf",
                    delete=False,   # must be False on Windows
                )
                try:
                    tmp.write(uf.read())
                    tmp.flush()
                    tmp.close()                        # close before pdfplumber reads

                    rows = parse_shipping_bill(tmp.name)
                    for r in rows:
                        r["Source_File"] = fname
                    all_rows.extend(rows)
                finally:
                    try:
                        Path(tmp.name).unlink(missing_ok=True)   # always delete
                    except Exception:
                        pass

                with log_container:
                    st.markdown(
                        f"&nbsp;&nbsp;✅ `{fname}` → **{len(rows)} row(s)**"
                    )

            except Exception as exc:
                file_errors.append((fname, str(exc)))
                with log_container:
                    st.markdown(f"&nbsp;&nbsp;❌ `{fname}` — {exc}")
                traceback.print_exc()

        # ── Final progress ────────────────────────────────────
        progress_bar.progress(1.0, text="Done!")
        status_text.empty()

        st.markdown("---")

        # ── Stats ─────────────────────────────────────────────
        col1, col2, col3 = st.columns(3)
        col1.metric("Files Processed", total)
        col2.metric("Rows Extracted",  len(all_rows))
        col3.metric("Errors",          len(file_errors))

        # ── Preview table ──────────────────────────────────────
        if all_rows:
            st.markdown("#### 🔍 Preview (first 20 rows)")
            df_preview = pd.DataFrame(all_rows)
            for col in COLUMN_ORDER:
                if col not in df_preview.columns:
                    df_preview[col] = None
            df_preview = df_preview[COLUMN_ORDER]
            st.dataframe(df_preview.head(20), use_container_width=True)

            # ── Build Excel in memory (no disk write) ─────────
            st.markdown("#### ⬇️ Download")
            excel_bytes = build_excel_bytes(all_rows)

            st.download_button(
                label="⬇️  Download Excel  (.xlsx)",
                data=excel_bytes,
                file_name="shipping_bills_extracted.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.error("No data could be extracted from the uploaded files.", icon="❌")

        # ── Show errors if any ────────────────────────────────
        if file_errors:
            with st.expander(f"⚠️  {len(file_errors)} file(s) had errors — click to view"):
                for fname, err in file_errors:
                    st.markdown(f"**`{fname}`** — {err}")

else:
    st.info(
        "👆  Upload one or more Shipping Bill PDFs above to begin extraction.",
        icon="ℹ️",
    )

# ── Footer ────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    "<p style='text-align:center;color:#475569;font-size:.78rem;'>"
    "Indian Customs Shipping Bill Extractor &nbsp;·&nbsp; "
    "Pure PDF Parsing &nbsp;·&nbsp; No External API &nbsp;·&nbsp; "
    "All files processed in memory only</p>",
    unsafe_allow_html=True,
)
